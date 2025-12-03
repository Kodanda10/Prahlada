from openpyxl import load_workbook
from pathlib import Path

def inspect_excel():
    path = Path("LGD_Chhattisgarh Complete Geographical Data/District_Subdistrict_Village_Gps_2025-11-27_01-45-36.xlsx")
    
    print("📖 Reading Excel header with openpyxl...")
    wb = load_workbook(filename=path, read_only=True)
    ws = wb.active
    
    print("\nColumns:")
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for col in row:
            print(f" - {col}")
            
    print("\nSample Row:")
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        print(row)

if __name__ == "__main__":
    inspect_excel()
